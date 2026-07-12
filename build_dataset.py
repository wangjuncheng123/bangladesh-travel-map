import csv
import difflib
import json
import re
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT = Path(__file__).parent / "data"
OFFICIAL = "https://beautifulbangladesh.gov.bd/"
DIVISION_BOUNDS = {
    "Dhaka": (22.7, 25.3, 89.2, 91.35),
    "Chattogram": (20.4, 24.4, 90.4, 92.8),
    "Khulna": (21.4, 24.4, 88.3, 90.8),
    "Rajshahi": (23.4, 25.7, 88.0, 90.9),
    "Rangpur": (24.7, 26.8, 88.0, 90.8),
    "Sylhet": (23.8, 25.5, 90.7, 92.6),
    "Barishal": (21.7, 23.7, 89.5, 91.2),
    "Mymensingh": (23.8, 25.7, 89.2, 91.8),
}

# Publicly listed tourism attractions, normalized for map use. Administrative
# fields are deliberately conservative; uncertain upazila values are blank.
SEEDS = [
    ("Sundarbans", "Mangrove forest / World Heritage", "Khulna", "Khulna, Bagerhat, Satkhira", ""),
    ("Sixty Dome Mosque", "Mosque / World Heritage", "Khulna", "Bagerhat", "Bagerhat Sadar"),
    ("Khan Jahan Ali's Tomb", "Historic mausoleum", "Khulna", "Bagerhat", "Bagerhat Sadar"),
    ("Shilaidaha Kuthibari", "Historic house / museum", "Khulna", "Kushtia", "Kumarkhali"),
    ("Lalon Shah's Shrine", "Cultural heritage", "Khulna", "Kushtia", "Kumarkhali"),
    ("Mujibnagar Memorial Complex", "Memorial", "Khulna", "Meherpur", "Mujibnagar"),
    ("Michael Madhusudan Dutt Memorial House", "Historic house / museum", "Khulna", "Jashore", "Keshabpur"),
    ("Hardinge Bridge", "Historic bridge", "Rajshahi, Khulna", "Pabna, Kushtia", "Ishwardi, Bheramara"),
    ("Paharpur Buddhist Monastery", "Archaeological site / World Heritage", "Rajshahi", "Naogaon", "Badalgachhi"),
    ("Mahasthangarh", "Archaeological site", "Rajshahi", "Bogura", "Shibganj"),
    ("Puthia Temple Complex", "Temple complex", "Rajshahi", "Rajshahi", "Puthia"),
    ("Puthia Rajbari", "Palace", "Rajshahi", "Rajshahi", "Puthia"),
    ("Bagha Mosque", "Historic mosque", "Rajshahi", "Rajshahi", "Bagha"),
    ("Choto Sona Mosque", "Historic mosque", "Rajshahi", "Chapainawabganj", "Shibganj"),
    ("Kusumba Mosque", "Historic mosque", "Rajshahi", "Naogaon", "Manda"),
    ("Varendra Research Museum", "Museum", "Rajshahi", "Rajshahi", "Boalia"),
    ("Natore Rajbari", "Palace", "Rajshahi", "Natore", "Natore Sadar"),
    ("Uttara Ganabhaban", "Palace / heritage", "Rajshahi", "Natore", "Natore Sadar"),
    ("Rabindra Kachharibari, Patisar", "Historic house", "Rajshahi", "Naogaon", "Atrai"),
    ("Chalan Beel", "Wetland", "Rajshahi", "Natore, Pabna, Sirajganj", ""),
    ("Kantajew Temple", "Historic temple", "Rangpur", "Dinajpur", "Kaharole"),
    ("Ramsagar National Park", "National park / reservoir", "Rangpur", "Dinajpur", "Dinajpur Sadar"),
    ("Tajhat Palace", "Palace / museum", "Rangpur", "Rangpur", "Rangpur Sadar"),
    ("Nilsagar", "Lake", "Rangpur", "Nilphamari", "Nilphamari Sadar"),
    ("Teesta Barrage", "Barrage / viewpoint", "Rangpur", "Lalmonirhat, Nilphamari", "Hatibandha, Dimla"),
    ("Bhitargarh", "Archaeological site", "Rangpur", "Panchagarh", "Panchagarh Sadar"),
    ("Tetulia Dak Bungalow and Picnic Corner", "Landscape / viewpoint", "Rangpur", "Panchagarh", "Tetulia"),
    ("Sitakot Vihara", "Archaeological site", "Rangpur", "Dinajpur", "Nawabganj"),
    ("Begum Rokeya Memorial", "Memorial / museum", "Rangpur", "Rangpur", "Mithapukur"),
    ("Lalbagh Fort", "Historic fort", "Dhaka", "Dhaka", "Lalbagh"),
    ("Ahsan Manzil", "Palace / museum", "Dhaka", "Dhaka", "Kotwali"),
    ("Bangladesh National Parliament House", "Modern architecture", "Dhaka", "Dhaka", "Sher-e-Bangla Nagar"),
    ("National Martyrs' Memorial", "National memorial", "Dhaka", "Dhaka", "Savar"),
    ("Central Shaheed Minar", "National memorial", "Dhaka", "Dhaka", "Shahbag"),
    ("Bangladesh National Museum", "Museum", "Dhaka", "Dhaka", "Shahbag"),
    ("Liberation War Museum", "Museum", "Dhaka", "Dhaka", "Sher-e-Bangla Nagar"),
    ("Dhakeshwari Temple", "Temple", "Dhaka", "Dhaka", "Lalbagh"),
    ("Baitul Mukarram", "National mosque", "Dhaka", "Dhaka", "Paltan"),
    ("Armenian Church, Dhaka", "Historic church", "Dhaka", "Dhaka", "Bangshal"),
    ("Star Mosque", "Historic mosque", "Dhaka", "Dhaka", "Bangshal"),
    ("Curzon Hall", "Historic architecture", "Dhaka", "Dhaka", "Shahbag"),
    ("Suhrawardy Udyan", "Historic park", "Dhaka", "Dhaka", "Shahbag"),
    ("Sonargaon", "Historic city", "Dhaka", "Narayanganj", "Sonargaon"),
    ("Panam City", "Historic urban site", "Dhaka", "Narayanganj", "Sonargaon"),
    ("Goaldi Mosque", "Historic mosque", "Dhaka", "Narayanganj", "Sonargaon"),
    ("Baliati Palace", "Palace", "Dhaka", "Manikganj", "Saturia"),
    ("Idrakpur Fort", "Historic fort", "Dhaka", "Munshiganj", "Munshiganj Sadar"),
    ("Bhawal National Park", "National park", "Dhaka", "Gazipur", "Gazipur Sadar"),
    ("Madhupur National Park", "National park", "Dhaka", "Tangail", "Madhupur"),
    ("Atia Mosque", "Historic mosque", "Dhaka", "Tangail", "Delduar"),
    ("Wari-Bateshwar ruins", "Archaeological site", "Dhaka", "Narsingdi", "Belabo"),
    ("Cox's Bazar Beach", "Sea beach", "Chattogram", "Cox's Bazar", "Cox's Bazar Sadar"),
    ("Inani Beach", "Sea beach", "Chattogram", "Cox's Bazar", "Ukhia"),
    ("Himchari National Park", "National park / waterfall", "Chattogram", "Cox's Bazar", "Cox's Bazar Sadar, Ramu"),
    ("Saint Martin's Island", "Island / coral beach", "Chattogram", "Cox's Bazar", "Teknaf"),
    ("Chhera Island", "Island", "Chattogram", "Cox's Bazar", "Teknaf"),
    ("Sonadia Island", "Island / beach", "Chattogram", "Cox's Bazar", "Maheshkhali"),
    ("Maheshkhali Island", "Island / temple", "Chattogram", "Cox's Bazar", "Maheshkhali"),
    ("Patenga Beach", "Sea beach", "Chattogram", "Chattogram", "Patenga"),
    ("Foy's Lake", "Lake", "Chattogram", "Chattogram", "Khulshi"),
    ("Sitakunda Botanical Garden and Eco-Park", "Eco-park", "Chattogram", "Chattogram", "Sitakunda"),
    ("Chandranath Hill", "Hill / temple", "Chattogram", "Chattogram", "Sitakunda"),
    ("Khoiyachora Waterfall", "Waterfall", "Chattogram", "Chattogram", "Mirsharai"),
    ("Sajek Valley", "Hill landscape", "Chattogram", "Rangamati", "Baghaichhari"),
    ("Kaptai Lake", "Lake", "Chattogram", "Rangamati", "Kaptai"),
    ("Shuvolong Waterfall", "Waterfall", "Chattogram", "Rangamati", "Barkal"),
    ("Nilgiri, Bandarban", "Hill viewpoint", "Chattogram", "Bandarban", "Ruma"),
    ("Nilachal", "Hill viewpoint", "Chattogram", "Bandarban", "Bandarban Sadar"),
    ("Boga Lake", "Mountain lake", "Chattogram", "Bandarban", "Ruma"),
    ("Keokradong", "Mountain", "Chattogram", "Bandarban", "Ruma"),
    ("Nafakhum Waterfall", "Waterfall", "Chattogram", "Bandarban", "Thanchi"),
    ("Amiakhum Waterfall", "Waterfall", "Chattogram", "Bandarban", "Thanchi"),
    ("Buddha Dhatu Jadi", "Buddhist temple", "Chattogram", "Bandarban", "Bandarban Sadar"),
    ("Mainamati Buddhist ruins", "Archaeological site", "Chattogram", "Cumilla", "Cumilla Adarsha Sadar"),
    ("Shalban Vihara", "Archaeological site", "Chattogram", "Cumilla", "Cumilla Sadar Dakshin"),
    ("Mainamati War Cemetery", "War cemetery", "Chattogram", "Cumilla", "Cumilla Adarsha Sadar"),
    ("Nijhum Dwip", "Island / national park", "Chattogram", "Noakhali", "Hatiya"),
    ("Ratargul Swamp Forest", "Freshwater swamp forest", "Sylhet", "Sylhet", "Gowainghat"),
    ("Jaflong", "River / hill landscape", "Sylhet", "Sylhet", "Gowainghat"),
    ("Bichanakandi", "River / stone landscape", "Sylhet", "Sylhet", "Gowainghat"),
    ("Lalakhal", "River landscape", "Sylhet", "Sylhet", "Jaintiapur"),
    ("Shrine of Shah Jalal", "Religious shrine", "Sylhet", "Sylhet", "Sylhet Sadar"),
    ("Lawachara National Park", "National park", "Sylhet", "Moulvibazar", "Kamalganj"),
    ("Madhabkunda Waterfall", "Waterfall / eco-park", "Sylhet", "Moulvibazar", "Barlekha"),
    ("Madhabpur Lake", "Lake / tea estate", "Sylhet", "Moulvibazar", "Kamalganj"),
    ("Hakaluki Haor", "Wetland / haor", "Sylhet", "Moulvibazar, Sylhet", "Barlekha, Kulaura, Juri, Golapganj, Fenchuganj"),
    ("Baikka Beel", "Wetland sanctuary", "Sylhet", "Moulvibazar", "Sreemangal"),
    ("Hum Hum Waterfall", "Waterfall", "Sylhet", "Moulvibazar", "Kamalganj"),
    ("Satchari National Park", "National park", "Sylhet", "Habiganj", "Chunarughat"),
    ("Rema-Kalenga Wildlife Sanctuary", "Wildlife sanctuary", "Sylhet", "Habiganj", "Chunarughat"),
    ("Tanguar Haor", "Ramsar wetland", "Sylhet", "Sunamganj", "Tahirpur, Madhyanagar"),
    ("Kuakata Beach", "Sea beach", "Barishal", "Patuakhali", "Kalapara"),
    ("Fatrar Char", "Mangrove forest / beach", "Barishal", "Barguna", "Taltali"),
    ("Bhimruli Floating Guava Market", "River market / cultural landscape", "Barishal", "Jhalokati", "Jhalokati Sadar"),
    ("Durgasagar", "Lake", "Barishal", "Barishal", "Babuganj"),
    ("Oxford Mission Church", "Historic church", "Barishal", "Barishal", "Barishal Sadar"),
    ("Char Kukri-Mukri Wildlife Sanctuary", "Wildlife sanctuary / island", "Barishal", "Bhola", "Char Fasson"),
    ("Birishiri", "Cultural / hill landscape", "Mymensingh", "Netrokona", "Durgapur"),
    ("Someshwari River", "River landscape", "Mymensingh", "Netrokona", "Durgapur"),
    ("Muktagacha Palace", "Palace", "Mymensingh", "Mymensingh", "Muktagacha"),
    ("Shashi Lodge", "Palace / heritage", "Mymensingh", "Mymensingh", "Mymensingh Sadar"),
    ("Zainul Abedin Museum", "Museum", "Mymensingh", "Mymensingh", "Mymensingh Sadar"),
    ("Hajiganj Fort", "Historic fort / UNESCO tentative component", "Dhaka", "Narayanganj", "Narayanganj Sadar"),
    ("Sonakanda Fort", "Historic fort / UNESCO tentative component", "Dhaka", "Narayanganj", "Bandar"),
    ("Bharat Bhayna Buddhist Temple", "Archaeological site / UNESCO tentative component", "Khulna", "Jashore", "Keshabpur"),
    ("Bibi Chini Mosque", "Historic mosque / UNESCO tentative component", "Barishal", "Barguna", "Betagi"),
    ("Nayabad Mosque", "Historic mosque / UNESCO tentative component", "Rangpur", "Dinajpur", "Kaharole"),
    ("National Botanical Garden of Bangladesh", "Botanical garden", "Dhaka", "Dhaka", "Mirpur"),
    ("Bangabandhu Sheikh Mujib Safari Park, Gazipur", "Safari park", "Dhaka", "Gazipur", "Sreepur"),
    ("Dulahazra Safari Park", "Safari park", "Chattogram", "Cox's Bazar", "Chakaria"),
    ("Khadimnagar National Park", "National park", "Sylhet", "Sylhet", "Sylhet Sadar"),
    ("Kaptai National Park", "National park", "Chattogram", "Rangamati", "Kaptai"),
    ("Teknaf Wildlife Sanctuary", "Wildlife sanctuary", "Chattogram", "Cox's Bazar", "Teknaf"),
    ("Chunati Wildlife Sanctuary", "Wildlife sanctuary", "Chattogram", "Chattogram", "Lohagara"),
    ("Madhutila Eco Park", "Eco-park / hill landscape", "Mymensingh", "Sherpur", "Nalitabari"),
    ("Tea Gardens of Sreemangal", "Tea landscape", "Sylhet", "Moulvibazar", "Sreemangal"),
    ("Jadukata River", "River landscape", "Sylhet", "Sunamganj", "Tahirpur"),
    ("Sandwip Island", "Island / coastal landscape", "Chattogram", "Chattogram", "Sandwip"),
    ("Alutila Cave", "Natural cave / hill attraction", "Chattogram", "Khagrachhari", "Matiranga"),
    ("Satla Beel", "Wetland / water-lily landscape", "Barishal", "Barishal", "Uzirpur"),
    ("Padma River, Rajshahi", "River landscape", "Rajshahi", "Rajshahi", ""),
    ("S. M. Sultan Memorial Complex", "Art museum / cultural landmark", "Khulna", "Narail", "Narail Sadar"),
]

SPECIAL_SOURCES = {
    "Sundarbans": "https://whc.unesco.org/en/list/798 | https://rsis.ramsar.org/RISapp/files/RISrep/BD560RIS.pdf",
    "Sixty Dome Mosque": "https://whc.unesco.org/en/list/321",
    "Paharpur Buddhist Monastery": "https://whc.unesco.org/en/list/322",
    "Tanguar Haor": "https://rsis.ramsar.org/RISapp/files/RISrep/BD1031RIS.pdf",
    "Hajiganj Fort": "https://whc.unesco.org/en/tentativelists/6675/",
    "Sonakanda Fort": "https://whc.unesco.org/en/tentativelists/6675/",
    "Bharat Bhayna Buddhist Temple": "https://whc.unesco.org/en/tentativelists/6669/",
    "Bibi Chini Mosque": "https://whc.unesco.org/en/tentativelists/6672/",
    "Nayabad Mosque": "https://whc.unesco.org/en/tentativelists/6672/",
    "National Botanical Garden of Bangladesh": "https://sufal.bforest.gov.bd/redlist/wp-content/uploads/2024/09/Red-List-Summary-English.pdf",
    "Bangabandhu Sheikh Mujib Safari Park, Gazipur": "https://bfis.bforest.gov.bd/library/wp-content/uploads/2021/02/Atlas-Forest-Mamagement.pdf",
    "Dulahazra Safari Park": "https://bfis.bforest.gov.bd/library/wp-content/uploads/2021/02/Atlas-Forest-Mamagement.pdf",
    "Khadimnagar National Park": "https://sufal.bforest.gov.bd/redlist/wp-content/uploads/2024/09/Red-List-Summary-English.pdf",
    "Kaptai National Park": "https://sufal.bforest.gov.bd/redlist/wp-content/uploads/2024/09/Red-List-Summary-English.pdf",
    "Teknaf Wildlife Sanctuary": "https://sufal.bforest.gov.bd/redlist/wp-content/uploads/2024/09/Red-List-Summary-English.pdf",
    "Chunati Wildlife Sanctuary": "https://sufal.bforest.gov.bd/redlist/wp-content/uploads/2024/09/Red-List-Summary-English.pdf",
    "Madhutila Eco Park": "https://www.beautifulbangladesh.gov.bd/newsletter/single/468",
    "Tea Gardens of Sreemangal": "https://beautifulbangladesh.gov.bd/district-destination/moulvibazar/green-zone/24",
    "Jadukata River": "https://beautifulbangladesh.gov.bd/district-destination/sunamganj/land-of-rivers/26",
    "Sandwip Island": "https://beautifulbangladesh.gov.bd/district-destination/chattogram/island/16",
    "Alutila Cave": "https://beautifulbangladesh.gov.bd/district-destination/khagrachari/hill-tracts-waterfalls/125",
    "Satla Beel": "https://beautifulbangladesh.gov.bd/district-destination/barisal/land-of-rivers/138",
    "Padma River, Rajshahi": "https://beautifulbangladesh.gov.bd/district-destination/rajshahi/land-of-rivers/27",
    "S. M. Sultan Memorial Complex": "https://beautifulbangladesh.gov.bd/district-destination/narail/landmarks/233",
    "Hardinge Bridge": "https://www.supremecourt.gov.bd/resources/documents/2157381_WritPetitionNo.7030of2022.pdf",
    "Teesta Barrage": "https://www.lalmonirhat.gov.bd/en/site/tourist_spot/oWYp-%E0%A6%A4%E0%A6%BF%E0%A6%B8%E0%A7%8D%E0%A6%A4%E0%A6%BE-%E0%A6%AC%E0%A7%8D%E0%A6%AF%E0%A6%BE%E0%A6%B0%E0%A7%87%E0%A6%9C-%E0%A6%93-%E0%A6%85%E0%A6%AC%E0%A6%B8%E0%A6%B0--%E0%A6%B0%E0%A7%87%E0%A6%B8%E0%A7%8D%E0%A6%9F-%E0%A6%B9%E0%A6%BE%E0%A6%89%E0%A6%9C",
    "Liberation War Museum": "https://liberationwarmuseumbd.org/page/contact-us",
    "Himchari National Park": "https://bforest.gov.bd/pages/static-pages/6922dfd3933eb65569e242f5",
    "Nilgiri, Bandarban": "https://bhdc.gov.bd/site/page/bc5994c3-24fe-438f-91e1-c76dd05d367f/",
    "Fatrar Char": "https://taltali.barguna.gov.bd/pages/static-pages/6978a7a235ce18e1c0674ea1",
    "Bhimruli Floating Guava Market": "https://www.jhalakathi.gov.bd/en/site/tourist_spot/%E0%A6%AD%E0%A6%BF%E0%A6%AE%E0%A6%B0%E0%A7%81%E0%A6%B2%E0%A7%80-%E0%A6%AA%E0%A7%87%E0%A7%9F%E0%A6%BE%E0%A6%B0%E0%A6%BE-%E0%A6%AC%E0%A6%BE%E0%A6%97%E0%A6%BE%E0%A6%A8-%E0%A6%93-%E0%A6%AD%E0%A6%BE%E0%A6%B8%E0%A6%AE%E0%A6%BE%E0%A6%A8-%E0%A6%AA%E0%A7%87%E0%A7%9F%E0%A6%BE%E0%A6%B0%E0%A6%BE-%E0%A6%AC%E0%A6%BE%E0%A6%9C%E0%A6%BE%E0%A6%B0",
}

PARENT_SITES = {
    "Panam City": "Sonargaon", "Goaldi Mosque": "Sonargaon",
    "Puthia Temple Complex": "Puthia", "Puthia Rajbari": "Puthia",
    "Shalban Vihara": "Mainamati Buddhist ruins",
    "Chhera Island": "Saint Martin's Island",
    "Sixty Dome Mosque": "Historic Mosque City of Bagerhat",
}


def api_json(url):
    req = urllib.request.Request(url, headers={"User-Agent": "BD-map-research/1.0"})
    with urllib.request.urlopen(req, timeout=20) as response:
        return json.load(response)


def wikidata(name):
    params = urllib.parse.urlencode({"action": "wbsearchentities", "search": name, "language": "en", "format": "json", "limit": 8, "origin": "*"})
    found = api_json("https://www.wikidata.org/w/api.php?" + params).get("search", [])
    if not found:
        return {}
    ids = "|".join(x["id"] for x in found)
    p = urllib.parse.urlencode({"action": "wbgetentities", "ids": ids, "props": "labels|descriptions|claims|sitelinks", "languages": "en|bn|zh", "format": "json", "origin": "*"})
    entities = api_json("https://www.wikidata.org/w/api.php?" + p).get("entities", {})
    for candidate in found:
        entity = entities.get(candidate["id"], {})
        claims = entity.get("claims", {})
        coord = None
        for c in claims.get("P625", []):
            value = c.get("mainsnak", {}).get("datavalue", {}).get("value")
            if isinstance(value, dict) and "latitude" in value:
                coord = (value["latitude"], value["longitude"])
                break
        countries = [c.get("mainsnak", {}).get("datavalue", {}).get("value", {}).get("id") for c in claims.get("P17", [])]
        # Many landmarks omit P17. A coordinate inside Bangladesh's bounding
        # box is an additional conservative acceptance test for those items.
        in_bd_bbox = bool(coord and 20.5 <= coord[0] <= 26.7 and 88.0 <= coord[1] <= 93.0)
        if "Q902" not in countries and not in_bd_bbox:
            continue
        image = ""
        if claims.get("P18"):
            filename = claims["P18"][0].get("mainsnak", {}).get("datavalue", {}).get("value", "")
            if filename:
                image = "https://commons.wikimedia.org/wiki/Special:FilePath/" + urllib.parse.quote(filename.replace(" ", "_"))
        labels = entity.get("labels", {})
        desc = entity.get("descriptions", {}).get("en", {}).get("value", "")
        return {"id": candidate["id"], "lat": coord[0] if coord else "", "lon": coord[1] if coord else "", "en": labels.get("en", {}).get("value", name), "bn": labels.get("bn", {}).get("value", ""), "zh": labels.get("zh", {}).get("value", ""), "desc": desc, "image": image}
    return {}


def photon_geocode(name, district, division):
    query = f"{name}, {district}, Bangladesh"
    url = "https://photon.komoot.io/api/?" + urllib.parse.urlencode({"q": query, "limit": 5, "lang": "en"})
    data = api_json(url)
    def norm(value):
        value = re.sub(r"[^a-z0-9]+", " ", (value or "").lower())
        stop = {"the", "of", "and", "bangladesh"}
        return " ".join(x for x in value.split() if x not in stop)
    wanted_name = norm(name)
    wanted_district = norm(district.split(",")[0])
    wanted_divisions = [norm(x) for x in division.split(",")]
    for feature in data.get("features", []):
        coords = feature.get("geometry", {}).get("coordinates", [])
        props = feature.get("properties", {})
        found_name = norm(props.get("name", ""))
        admin_text = norm(" ".join(str(props.get(k, "")) for k in ("city", "county", "district", "state")))
        similarity = difflib.SequenceMatcher(None, wanted_name, found_name).ratio()
        name_ok = bool(found_name and (wanted_name in found_name or found_name in wanted_name or similarity >= 0.82))
        admin_ok = bool((wanted_district and wanted_district in admin_text) or any(x and x in admin_text for x in wanted_divisions))
        if len(coords) == 2 and name_ok and admin_ok and 88.0 <= coords[0] <= 93.0 and 20.5 <= coords[1] <= 26.7:
            return {"lat": coords[1], "lon": coords[0], "geocode_name": props.get("name", ""), "geocode_admin": admin_text}
    return {}


def best_season(category):
    lower = category.lower()
    if any(x in lower for x in ("waterfall", "river", "wetland", "haor")):
        return "雨季至雨季后（约6–11月）；出行前核实水位与安全情况"
    if any(x in lower for x in ("beach", "island", "mangrove", "national park", "forest", "mountain", "hill")):
        return "旱季（约11月至次年3月）"
    return "全年；通常11月至次年3月天气较舒适"


def category_group(category):
    c = category.lower()
    if any(x in c for x in ("mosque", "temple", "church", "shrine", "buddhist")): return "宗教与考古"
    if any(x in c for x in ("fort", "palace", "heritage", "historic", "archaeological", "world heritage")): return "历史文化"
    if any(x in c for x in ("museum", "memorial", "architecture", "cemetery")): return "博物馆与纪念地"
    if any(x in c for x in ("beach", "island", "coastal")): return "海滩与岛屿"
    if any(x in c for x in ("waterfall", "mountain", "hill", "cave", "viewpoint")): return "山地与瀑布"
    if any(x in c for x in ("wetland", "haor", "lake", "river", "reservoir", "barrage")): return "河流湖泊与湿地"
    if any(x in c for x in ("park", "forest", "sanctuary", "garden", "safari", "mangrove")): return "森林公园与保护区"
    return "城市与文化景观"


def safe_description(name, category, division, district):
    group = category_group(category)
    return f"{name}位于孟加拉国{division}大区的{district}，在本数据集中归为“{group}”。开放、通行和许可信息可能随季节或管理政策变化，出行前应向主管部门复核。"


def coordinate_in_division(lat, lon, division):
    if lat == "" or lon == "":
        return False
    for part in (x.strip() for x in division.split(",")):
        south, north, west, east = DIVISION_BOUNDS[part]
        if south <= float(lat) <= north and west <= float(lon) <= east:
            return True
    return False


def main():
    OUT.mkdir(exist_ok=True)
    cache_path = OUT / "wikidata_cache.json"
    cache = json.loads(cache_path.read_text()) if cache_path.exists() else {}
    rows = []
    for index, (name, category, division, district, upazila) in enumerate(SEEDS, 1):
        if name not in cache:
            try:
                cache[name] = wikidata(name)
            except Exception as exc:
                cache[name] = {"error": str(exc)}
            time.sleep(0.12)
        wd = cache[name]
        if wd.get("lat", "") != "" and not coordinate_in_division(wd["lat"], wd["lon"], division):
            wd["rejected_coordinate"] = [wd.get("lat"), wd.get("lon")]
            wd["lat"] = ""; wd["lon"] = ""; wd.pop("geocoded", None)
        if not wd.get("lat") and not wd.get("photon_checked"):
            try:
                geo = photon_geocode(name, district.split(",")[0], division)
                if geo:
                    wd.update(geo)
                    wd["geocoded"] = True
            except Exception:
                pass
            wd["photon_checked"] = True
            time.sleep(0.12)
        source = OFFICIAL
        source_type = "Bangladesh Tourism Board directory"
        status = "旅游局目录匹配"
        if name in SPECIAL_SOURCES:
            source += " | " + SPECIAL_SOURCES[name]
            source_type += " | specific authoritative source"
            status = "景点级权威来源已核实"
        if wd.get("id"):
            source += " | https://www.wikidata.org/wiki/" + wd["id"]
            source_type += " | Wikidata"
            if wd.get("lat") != "" and not wd.get("geocoded"):
                status += "；Wikidata坐标自动匹配（待人工复核）"
        else:
            if wd.get("lat", "") == "":
                status += "；坐标待补"
        if wd.get("geocoded"):
            source += f" | https://www.openstreetmap.org/?mlat={wd['lat']}&mlon={wd['lon']}#map=15/{wd['lat']}/{wd['lon']}"
            source_type += " | OpenStreetMap Photon"
            status += "；OSM坐标自动匹配（待人工复核）"
        lat = f"{float(wd['lat']):.7f}" if wd.get("lat", "") != "" else ""
        lon = f"{float(wd['lon']):.7f}" if wd.get("lon", "") != "" else ""
        rows.append({
            "id": f"BD-POI-{index:04d}", "name_zh": wd.get("zh", ""), "name_en": wd.get("en", name), "name_bn": wd.get("bn", ""),
            "category_group": category_group(category), "category": category, "parent_site": PARENT_SITES.get(name, ""),
            "division": division, "district": district, "upazila_or_thana": upazila,
            "latitude": lat, "longitude": lon, "address": "",
            "description_zh": safe_description(name, category, division, district),
            "opening_hours": "", "ticket_price": "", "best_season": best_season(category),
            "transport": "出发前通过当地旅游部门或地图服务核实路线", "contact": "", "website": "",
            "image_url": wd.get("image", ""), "source_urls": source, "source_type": source_type, "verification_status": status,
            "last_verified": date.today().isoformat(), "notes": "开放时间、门票和通行许可具有时效性，未获可靠来源时留空。",
        })
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2))
    csv_path = OUT / "bangladesh_attractions.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader(); writer.writerows(rows)
    wb = Workbook(); ws = wb.active; ws.title = "景点数据"
    headers = list(rows[0]); ws.append(headers)
    for row in rows: ws.append([row[h] for h in headers])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]: cell.font = Font(color="FFFFFF", bold=True); cell.fill = fill; cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    wide = {"description_zh":48, "source_urls":65, "notes":48, "transport":36, "best_season":34, "image_url":45}
    medium = {"name_zh", "name_en", "name_bn", "category", "parent_site", "district", "upazila_or_thana", "address", "website", "source_type", "verification_status"}
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = wide.get(header, 25 if header in medium else 16)
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    meta = wb.create_sheet("字段说明")
    meta.append(["字段", "说明"])
    explanations = {
        "id":"项目内唯一编号", "name_zh/name_en/name_bn":"中文、英文、孟加拉文名称；缺少可靠标签时留空",
        "category_group":"供地图筛选使用的统一一级分类", "category":"更具体的景点类型", "parent_site":"上位景区或遗址群；用于处理父子景点",
        "division/district/upazila_or_thana":"行政区层级；跨区对象以逗号分隔", "latitude/longitude":"WGS84坐标；自动匹配坐标均在状态列标为待人工复核",
        "address":"详细地址；无可靠景点级来源时留空", "description_zh":"基于已核实位置和类别生成的简要中文说明",
        "opening_hours/ticket_price":"时效字段；无可靠当前资料时留空", "best_season":"按景观类型给出的通用建议，并非实时天气信息",
        "transport":"交通核实提示", "contact/website":"主管方联系方式和官网；未知时留空", "image_url":"Wikimedia Commons图片直链（如有）",
        "source_urls":"多个来源以竖线分隔", "source_type":"来源类别", "verification_status":"证据强度、自动匹配和待补项",
        "last_verified":"本次联网核实日期", "notes":"时效性和使用限制说明"
    }
    for key, value in explanations.items(): meta.append([key, value])
    meta.column_dimensions["A"].width = 34; meta.column_dimensions["B"].width = 80
    scope = wb.create_sheet("范围与限制")
    scope_rows = [
        ["项目", "说明"],
        ["数据范围", f"公开可核实的孟加拉国重点旅游景点与自然/文化保护地，共{len(rows)}条；不是法律意义上的穷尽名录。"],
        ["主要来源", "Bangladesh Tourism Board、UNESCO World Heritage Centre、Bangladesh Forest Department、Ramsar Sites Information Service、Wikidata、OpenStreetMap。"],
        ["坐标规则", "名称相似度与Division/District一致性校验后写入；自动坐标仍需项目上线前人工抽查。"],
        ["留空规则", "门票、开放时间、电话、官网等变化快；未找到可靠当前来源时留空，不猜测。"],
        ["更新建议", "门票与开放时间每月或上线前更新；保护区许可、山区安全和季节性通行应在出发前复核。"],
    ]
    for row in scope_rows: scope.append(row)
    scope.column_dimensions["A"].width = 22; scope.column_dimensions["B"].width = 110
    for row in scope.iter_rows():
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    xlsx_path = OUT / "bangladesh_attractions.xlsx"; wb.save(xlsx_path)
    print(json.dumps({"rows": len(rows), "csv": str(csv_path), "xlsx": str(xlsx_path), "with_coordinates": sum(bool(r["latitude"]) for r in rows), "specific_source_rows": sum("specific authoritative source" in r["source_type"] for r in rows)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
